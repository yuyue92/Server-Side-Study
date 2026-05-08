#!/usr/bin/env python3
"""
生成示例 Excel 配置文件 sample-config.xlsx
运行: python3 gen_sample_config.py
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def gen_with_openpyxl():
    wb = openpyxl.Workbook()
    
    HEADERS = ['name', '寄存器地址', '小数位数', '单位', '读写权限', '组件类型', '选项数据']
    
    sheets = {
        '电机参数': [
            ['电机转速',    40001, 1, 'RPM',  'READONLY',  'INPUT', ''],
            ['运行电流',    40002, 2, 'A',    'READONLY',  'INPUT', ''],
            ['母线电压',    40003, 1, 'V',    'READONLY',  'INPUT', ''],
            ['输出功率',    40004, 2, 'kW',   'READONLY',  'INPUT', ''],
            ['目标转速',    40005, 1, 'RPM',  'READWRITE', 'INPUT', ''],
            ['加速时间',    40006, 1, 's',    'READWRITE', 'INPUT', ''],
            ['减速时间',    40007, 1, 's',    'READWRITE', 'INPUT', ''],
            ['运行状态',    40008, 0, '',     'READONLY',  'SELECT', '[{"label":"停机","value":"0"},{"label":"运行","value":"1"},{"label":"故障","value":"2"}]'],
            ['控制模式',    40009, 0, '',     'READWRITE', 'SELECT', '[{"label":"速度控制","value":"0"},{"label":"转矩控制","value":"1"},{"label":"位置控制","value":"2"}]'],
            ['散热器温度',  40010, 1, '°C',   'READONLY',  'INPUT', ''],
            ['故障代码',    40011, 0, '',     'READONLY',  'INPUT', ''],
            ['累计运行时间',40012, 0, 'h',    'READONLY',  'INPUT', ''],
        ],
        '变频器参数': [
            ['载波频率',    40029, 1, 'kHz',  'READWRITE', 'INPUT', ''],
            ['最高频率',    40030, 2, 'Hz',   'READWRITE', 'INPUT', ''],
            ['最低频率',    40031, 2, 'Hz',   'READWRITE', 'INPUT', ''],
            ['额定电压',    40032, 0, 'V',    'READONLY',  'INPUT', ''],
            ['额定电流',    40033, 1, 'A',    'READONLY',  'INPUT', ''],
            ['电机极对数',  40034, 0, 'p',    'READWRITE', 'INPUT', ''],
            ['过载保护',    40035, 0, '%',    'READWRITE', 'INPUT', ''],
            ['启动方式',    40036, 0, '',     'READWRITE', 'SELECT', '[{"label":"直接启动","value":"0"},{"label":"软启动","value":"1"},{"label":"飞车启动","value":"2"}]'],
            ['制动方式',    40037, 0, '',     'READWRITE', 'SELECT', '[{"label":"自由停车","value":"0"},{"label":"减速停车","value":"1"},{"label":"直流制动","value":"2"}]'],
            ['PID比例增益', 40038, 2, '',     'READWRITE', 'INPUT', ''],
            ['PID积分时间', 40039, 2, 's',    'READWRITE', 'INPUT', ''],
        ],
        '保护参数': [
            ['过压保护值',  40050, 0, 'V',    'READWRITE', 'INPUT', ''],
            ['欠压保护值',  40051, 0, 'V',    'READWRITE', 'INPUT', ''],
            ['过流保护值',  40052, 1, 'A',    'READWRITE', 'INPUT', ''],
            ['过温保护值',  40053, 0, '°C',   'READWRITE', 'INPUT', ''],
            ['接地故障保护',40054, 0, '',     'READWRITE', 'SELECT', '[{"label":"禁用","value":"0"},{"label":"启用","value":"1"}]'],
            ['缺相保护',    40055, 0, '',     'READWRITE', 'SELECT', '[{"label":"禁用","value":"0"},{"label":"启用","value":"1"}]'],
        ],
    }
    
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        
        # Header style
        header_fill = PatternFill(start_color='1C2028', end_color='1C2028', fill_type='solid')
        header_font = Font(name='Consolas', bold=True, color='00D4AA')
        
        ws.append(HEADERS)
        for col in range(1, len(HEADERS)+1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row in rows:
            ws.append(row)
        
        # Column widths
        widths = [14, 12, 8, 6, 12, 10, 60]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    wb.save('sample-config.xlsx')
    print('✓ 已生成 sample-config.xlsx')

def gen_fallback():
    """不依赖 openpyxl，用 xlsxwriter 或 csv"""
    try:
        import xlsxwriter
        wb = xlsxwriter.Workbook('sample-config.xlsx')
        # minimal
        headers = ['name', '寄存器地址', '小数位数', '单位', '读写权限', '组件类型', '选项数据']
        ws = wb.add_worksheet('电机参数')
        for c, h in enumerate(headers):
            ws.write(0, c, h)
        ws.write(1, 0, '电机转速'); ws.write(1, 1, 40001); ws.write(1, 2, 1)
        ws.write(1, 3, 'RPM'); ws.write(1, 4, 'READONLY'); ws.write(1, 5, 'INPUT')
        wb.close()
        print('✓ 已生成 sample-config.xlsx (xlsxwriter)')
    except ImportError:
        print('请先安装: pip install openpyxl  或  pip install xlsxwriter')

if HAS_OPENPYXL:
    gen_with_openpyxl()
else:
    gen_fallback()
